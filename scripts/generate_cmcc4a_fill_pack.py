#!/usr/bin/env python
# -*- coding: utf-8 -*-

import json
import pathlib
import collections

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = pathlib.Path(r"C:\Users\chaoJ\Desktop\UnifiedModel")
DATA_DIR = ROOT / "data"
OUTPUTS_DIR = ROOT / "outputs"
DESKTOP_DIR = pathlib.Path(r"C:\Users\chaoJ\Desktop")


def iter_json_array_records(fp: pathlib.Path, limit: int | None = None):
    count = 0
    with fp.open("r", encoding="utf-8") as f:
        for raw in f:
            line = raw.strip()
            if not line or line in ("[", "]"):
                continue
            if line.endswith(","):
                line = line[:-1]
            try:
                obj = json.loads(line)
            except Exception:
                continue
            yield obj
            count += 1
            if limit and count >= limit:
                break


def autosize(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 48)


def add_sheet(wb, title, headers, rows):
    ws = wb.create_sheet(title)
    ws.append(headers)
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    autosize(ws)
    return ws


def load_expert_rows():
    expert_path = list(DESKTOP_DIR.glob("*MModel*采集表.xlsx"))[0]
    wb_exp = openpyxl.load_workbook(str(expert_path), data_only=True)
    main_ws = wb_exp.worksheets[1]
    case_ws = wb_exp.worksheets[2]

    expert_rows = []
    for r in range(6, main_ws.max_row + 1):
        vals = [main_ws.cell(r, c).value for c in range(1, 17)]
        if any(v not in (None, "") for v in vals):
            expert_rows.append({"row": r, "vals": vals})

    case_rows = []
    for r in range(6, case_ws.max_row + 1):
        vals = [case_ws.cell(r, c).value for c in range(1, 17)]
        if any(v not in (None, "") for v in vals):
            case_rows.append({"row": r, "vals": vals})

    return expert_path, expert_rows, case_rows


def build_profiles():
    profile_rows = []
    service_inventory = []
    source_stats = {}

    field_targets = {
        "log": [
            "serviceName",
            "traceId",
            "resource.attributes.service@name",
            "resource.attributes.host@ip",
            "log.attributes.log@file@path",
            "log.attributes.class@name",
            "log.attributes.thread@name",
            "time",
            "body",
        ],
        "metric": [
            "serviceName",
            "name",
            "resource.attributes.service@instance@id",
            "metric.attributes.topic",
            "metric.attributes.group",
            "metric.attributes.clientId",
            "metric.attributes.backend",
            "metric.attributes.url",
            "metric.attributes.status",
            "metric.attributes.server",
            "metric.attributes.env",
            "metric.attributes.nodename",
            "time",
            "value",
        ],
        "trace": [
            "serviceName",
            "traceId",
            "spanId",
            "parentSpanId",
            "name",
            "resource.attributes.service@instance@id",
            "resource.attributes.http@url",
            "span.attributes.db@type",
            "span.attributes.db@instance",
            "span.attributes.db@statement",
            "span.attributes.http@method",
            "status.code",
            "status.message",
            "startTime",
            "durationInNanos",
        ],
    }

    for kind, pattern, sample_limit in [
        ("log", "log/*.json", 120000),
        ("metric", "metric/*.json", 80000),
        ("trace", "trace/*.json", 120000),
    ]:
        field_freq = collections.Counter()
        field_example = {}
        service_ctr = collections.Counter()
        total = 0
        files = list(DATA_DIR.glob(pattern))
        per_file_limit = max(10000, sample_limit // max(1, len(files)))

        for fp in sorted(files):
            for obj in iter_json_array_records(fp, limit=per_file_limit):
                total += 1
                svc = obj.get("serviceName") or obj.get("resource.attributes.service@name")
                if svc:
                    service_ctr[svc] += 1
                for k in field_targets[kind]:
                    if k in obj and obj[k] not in (None, ""):
                        field_freq[k] += 1
                        field_example.setdefault(k, str(obj[k])[:120])

        source_stats[kind] = {
            "sampled_records": total,
            "services": service_ctr,
            "field_freq": field_freq,
            "field_example": field_example,
        }

        for field in field_targets[kind]:
            if field not in field_freq:
                continue

            use = ""
            target = ""
            if field in ("serviceName", "resource.attributes.service@name"):
                use, target = "实体主键 / 关联键", "cmcc4a.service"
            elif field == "traceId":
                use, target = "证据串联键", "跨 log/trace 关联"
            elif field in ("spanId", "parentSpanId"):
                use, target = "调用链拓扑键", "cmcc4a.trace.apm"
            elif "instance" in field or field == "resource.attributes.host@ip":
                use, target = "实例键", "中间件 / 数据库 / 网关实例"
            elif field.startswith("metric.attributes.topic") or field.startswith("metric.attributes.group"):
                use, target = "MQ 主键候选", "cmcc4a.mq"
            elif field.startswith("metric.attributes.url") or field.startswith("metric.attributes.backend"):
                use, target = "网关路由键", "cmcc4a.gateway"
            elif field.startswith("span.attributes.db@"):
                use, target = "依赖证据字段", "cmcc4a.database / cmcc4a.redis"
            else:
                use, target = "证据字段", ""

            confidence = "high" if field_freq[field] / max(total, 1) > 0.6 else "medium"
            profile_rows.append([
                kind,
                field,
                field_freq[field],
                field_example.get(field, ""),
                use,
                target,
                confidence,
                "",
            ])

        for svc, cnt in service_ctr.most_common(30):
            guessed_type = "service"
            svc_lower = svc.lower()
            if "redis" in svc_lower:
                guessed_type = "redis"
            elif "mysql" in svc_lower or "db" in svc_lower:
                guessed_type = "database"
            elif "mq" in svc_lower or "kafka" in svc_lower or "rocket" in svc_lower:
                guessed_type = "mq"
            elif "nginx" in svc_lower or "gateway" in svc_lower or "api" in svc_lower:
                guessed_type = "gateway"
            service_inventory.append([kind, svc, cnt, guessed_type, "", ""])

    return profile_rows, service_inventory, source_stats


def build_workbook():
    expert_path, expert_rows, case_rows = load_expert_rows()
    profile_rows, service_inventory, source_stats = build_profiles()

    entity_rows = [
        ["P0", "cmcc4a", "cmcc4a.service", "4A服务", "认证 / 登录 / 鉴权 / 资源 / 金库等应用服务", "serviceName", "resource.attributes.service@name, appId, routeId", "logs+traces", "专家表第6-18行已经覆盖；真实观测数据里也已有大量 serviceName", "确认统一命名、中文别名，以及是否需要拆分子类", "高", "先做"],
        ["P0", "cmcc4a", "cmcc4a.account", "账号/用户", "登录账号、鉴权账号、从账号、资源账号", "userId or accountId", "clientId, phone, tenantId", "logs", "专家多次提到 userId/accountId；观测字段仍需进一步确认", "确认账号是否作为核心实体，以及主键策略", "中", "候选"],
        ["P0", "cmcc4a", "cmcc4a.redis", "Redis/会话缓存", "Token 缓存、Session 缓存、事务锁缓存", "clusterId or instanceId", "keyPrefix, endpoint, serviceName", "metrics+traces+logs", "专家表第9/12行覆盖；metric 里有 redis 实例 ID；trace 里有 Redis 调用", "确认最终主键按 cluster 还是 instance", "高", "先做"],
        ["P0", "cmcc4a", "cmcc4a.database", "数据库", "业务库、磐维数据库、MySQL/PG 实例", "dbName or instanceId", "clusterId, sqlId, schema", "metrics+traces+logs", "专家表第6/12/17行及多个案例覆盖；trace 里有 db instance/statement", "确认数据库命名策略和实例层级", "高", "先做"],
        ["P1", "cmcc4a", "cmcc4a.mq", "消息队列/任务", "MQ Topic、消费者组、调度任务、补偿任务", "topic or consumerGroup or jobName", "taskId, clusterId, clientId", "metrics+logs", "专家表第10/12行覆盖；metric 里有 topic/group/clientId", "确认实体按 topic、consumerGroup 还是 job 建模", "中高", "先做"],
        ["P1", "cmcc4a", "cmcc4a.gateway", "网关/负载均衡", "入口网关、Nginx、路由、VIP", "gatewayName or routeId or vip", "listener, upstreamService, backend, url", "metrics+logs+traces", "专家表第11/18行覆盖；metric 里有 nginx/url/backend", "确认按网关实例还是按 route 建模", "中高", "先做"],
        ["P1", "cmcc4a", "cmcc4a.business_system", "业务系统/渠道", "门户、移动办公、省侧系统、接入应用", "systemName or appId", "channel, tenant", "experts+logs", "专家表多次提到调用方和受影响系统", "确认是否需要单独建实体集", "中", "候选"],
    ]

    relation_rows = [
        ["P0", "cmcc4a.business_system_calls_cmcc4a.service", "cmcc4a.business_system", "cmcc4a.service", "calls", "依据专家表第6/8/11行", "systemName/appId -> serviceName", "确认业务系统命名和边界", "用于回答哪个渠道调用哪个4A服务"],
        ["P0", "cmcc4a.service_use_cmcc4a.redis", "cmcc4a.service", "cmcc4a.redis", "use", "依据专家表第6/8/9/12行；trace 中已有 Redis 调用", "serviceName -> serviceName or endpoint mapping", "确认 redis 实体主键", "支撑认证/鉴权/短信对缓存的依赖分析"],
        ["P0", "cmcc4a.service_use_cmcc4a.database", "cmcc4a.service", "cmcc4a.database", "use", "依据专家表第6/12/17行；trace 中已有 db instance/statement", "serviceName -> db.instance/dbName", "确认 db 主键层级", "支撑慢 SQL 和数据库故障影响分析"],
        ["P1", "cmcc4a.service_sends_to_cmcc4a.mq", "cmcc4a.service", "cmcc4a.mq", "sends_to", "依据专家表第10/12行；metric 中已有 topic/group", "serviceName -> topic/consumerGroup", "确认 MQ 按 topic 还是 consumerGroup 建模", "支撑同步延迟和消息堆积分析"],
        ["P1", "cmcc4a.gateway_serves_cmcc4a.service", "cmcc4a.gateway", "cmcc4a.service", "serves", "依据专家表第11/18行；metric 中已有 backend/url", "backend/url -> serviceName/routeId", "确认 upstream 映射", "支撑入口层到服务层跳转"],
        ["P1", "cmcc4a.account_access_cmcc4a.business_system", "cmcc4a.account", "cmcc4a.business_system", "access", "依据专家表第6/8/13/17行", "userId/accountId -> systemName/appId", "仅当账号成为核心实体时再做", "支撑受影响用户分析"],
    ]

    dataset_rows = [
        ["P0", "log_set", "cmcc4a", "cmcc4a.log.application", "应用服务日志", "serviceName, traceId, time, body, resource.attributes.host@ip", "", str(DATA_DIR / "log"), "ais-configure, iam-manage, ais-amc, ais-consumer, ais-application-service 等", "确认最终 dataset 命名，以及是否拆分服务日志和网关日志"],
        ["P0", "trace_set", "cmcc4a", "cmcc4a.trace.apm", "APM 调用链", "serviceName, traceId, spanId, parentSpanId, name, startTime, durationInNanos, status.code", "trace_id_field=traceId; span_id_field=spanId; parent_span_id_field=parentSpanId; protocol=opentelemetry", str(DATA_DIR / "trace"), "ais-configure, iam-manage, ais-amc, ais-consumer 等", "确认是否拆成服务 trace 和网关 trace"],
        ["P0", "metric_set", "cmcc4a", "cmcc4a.metric.redis", "Redis 指标", "serviceName, resource.attributes.service@instance@id, name, value, time", "labels: instanceId", str(DATA_DIR / "metric"), "redisHcZ4a, redisAudit", "确认 Redis clusterId 来源"],
        ["P0", "metric_set", "cmcc4a", "cmcc4a.metric.database", "数据库指标", "serviceName, resource.attributes.service@instance@id, name, value, time", "labels: instanceId, env, nodename, type", str(DATA_DIR / "metric"), "EJpanweiDB, HCpanweiDB, z4amysql, skywalkingmysql", "确认 dbName/clusterId 口径"],
        ["P1", "metric_set", "cmcc4a", "cmcc4a.metric.mq", "MQ 指标", "serviceName, name, metric.attributes.topic, metric.attributes.group, metric.attributes.clientId, value, time", "labels: topic, group, clientId, cmd", str(DATA_DIR / "metric"), "rocketmq_exporter, logKafka2, csc-cm-it-mq", "确认 topic 和消费者组规范"],
        ["P1", "metric_set", "cmcc4a", "cmcc4a.metric.gateway", "网关/Nginx 指标", "serviceName, name, metric.attributes.url, metric.attributes.status, metric.attributes.backend, value, time", "labels: url, status, backend, site, server", str(DATA_DIR / "metric"), "LoadbalancerNginx, LoadbalancerNginxUpstream, apiNginx, webNginx, ngLogPush", "确认 gatewayName/routeId 来源"],
    ]

    datalink_rows = [
        ["cmcc4a.service", "cmcc4a.log.application", "related_to", "serviceName -> serviceName", "P0", "已可做", "服务到日志的关联最直接"],
        ["cmcc4a.service", "cmcc4a.trace.apm", "related_to", "serviceName -> serviceName", "P0", "已可做", "按服务名关联 trace 的覆盖率较高"],
        ["cmcc4a.service", "cmcc4a.log.application", "related_to", "traceId -> traceId / log.attributes.trace_id", "P0", "已可做", "用于把日志和 trace 串起来"],
        ["cmcc4a.redis", "cmcc4a.metric.redis", "related_to", "instanceId -> resource.attributes.service@instance@id", "P0", "待确认主键", "实例级指标已经存在"],
        ["cmcc4a.database", "cmcc4a.metric.database", "related_to", "instanceId -> resource.attributes.service@instance@id", "P0", "待确认主键", "trace 中也已有 db instance 字段可补强"],
        ["cmcc4a.database", "cmcc4a.trace.apm", "related_to", "dbName/instanceId -> span.attributes.db@instance", "P0", "待确认清洗规则", "用于从服务慢跳到 DB 慢"],
        ["cmcc4a.mq", "cmcc4a.metric.mq", "related_to", "topic -> metric.attributes.topic; consumerGroup -> metric.attributes.group", "P1", "可启动", "MQ 标签字段已存在"],
        ["cmcc4a.gateway", "cmcc4a.metric.gateway", "related_to", "routeId/url/backend -> metric.attributes.url/backend", "P1", "可启动", "入口层指标已经可见"],
        ["cmcc4a.account", "cmcc4a.log.application", "related_to", "userId/accountId -> body or structured fields", "P1", "待确认结构化字段", "账号实体暂时不要过早定型"],
    ]

    storage_rows = [
        ["external_storage", "cmcc4a.obs.log.local_export", "file-json-array", "4A 本地日志导出", str(DATA_DIR / "log"), "path=" + str(DATA_DIR / "log"), "后续可替换为 OpenSearch/SLS 索引信息"],
        ["external_storage", "cmcc4a.obs.metric.local_export", "file-json-array", "4A 本地指标导出", str(DATA_DIR / "metric"), "path=" + str(DATA_DIR / "metric"), "后续可替换为 Prom/OpenSearch"],
        ["external_storage", "cmcc4a.obs.trace.local_export", "file-json-array", "4A 本地调用链导出", str(DATA_DIR / "trace"), "path=" + str(DATA_DIR / "trace"), "后续可替换为 APM/Trace 后端"],
    ]

    storage_link_rows = [
        ["cmcc4a.log.application", "cmcc4a.obs.log.local_export", "按源 JSON 一比一映射", "后续补真实查询入口/索引"],
        ["cmcc4a.trace.apm", "cmcc4a.obs.trace.local_export", "traceId/spanId 等字段保持原样", "后续补真实查询入口/索引"],
        ["cmcc4a.metric.redis", "cmcc4a.obs.metric.local_export", "metric labels 保持原样", "后续补真实查询入口/索引"],
        ["cmcc4a.metric.database", "cmcc4a.obs.metric.local_export", "metric labels 保持原样", "后续补真实查询入口/索引"],
        ["cmcc4a.metric.mq", "cmcc4a.obs.metric.local_export", "metric labels 保持原样", "后续补真实查询入口/索引"],
        ["cmcc4a.metric.gateway", "cmcc4a.obs.metric.local_export", "metric labels 保持原样", "后续补真实查询入口/索引"],
    ]

    rule_rows = []
    for item in expert_rows:
        v = item["vals"]
        rule_rows.append([
            f"专家主填表!row{item['row']}",
            v[1],
            v[2],
            v[3],
            v[4],
            v[8],
            v[9],
            v[10],
            v[11],
            v[13],
            v[15],
        ])
    for item in case_rows:
        v = item["vals"]
        rule_rows.append([
            f"故障案例采集!row{item['row']}",
            v[1],
            v[3],
            v[6],
            v[7],
            v[8],
            "",
            v[12],
            v[6],
            v[10],
            v[14],
        ])

    open_questions = [
        ["Q1", "统一命名", "专家对象名和真实 serviceName 还没一一对齐，例如“认证服务/金库服务/GDC/self服务”分别对应哪些实际 serviceName 需要确认", "直接影响 entity_set 命名和 data_link 映射", "专家+工程"],
        ["Q2", "账号实体是否现在入模", "userId/accountId 在专家判断里很重要，但当前抽样观测数据里还没有明显稳定的结构化字段", "决定是否现在就落 cmcc4a.account", "专家+工程"],
        ["Q3", "Redis/DB/MQ 主键层级", "Redis、数据库、MQ 实体到底按 cluster、instance、dbName、topic、consumerGroup 还是 jobName 建模，需要统一粒度", "影响主键稳定性和关系稳定性", "工程+中间件专家"],
        ["Q4", "网关建模粒度", "是建 gateway 实体加 route 子实体，还是只保留 gateway + routeId 字段", "影响入口层分析深度", "工程"],
        ["Q5", "生产查询入口", "目前只知道本地导出路径，还缺生产上的 OpenSearch/Prom/APM 查询入口、索引名、时间字段", "影响 storage 和 storage_link 落地", "工程"],
        ["Q6", "示例内容清理", "专家表里仍有少量示例/模板内容，正式入模前需要删掉", "避免错误知识进入本体", "专家"],
    ]

    wb = Workbook()
    ws0 = wb.active
    ws0.title = "0_ReadMe"
    ws0.append(["Item", "Content"])
    for cell in ws0[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")

    readme_rows = [
        ["目标", "这份文件把 UnifiedModel 最小本体构建所需信息，整理成一份可直接补空的填报包。"],
        ["已结合的输入", f"1) 代码仓库 {ROOT} 里的 schema/example；2) 专家表 {expert_path.name}；3) data 目录下的真实 log/metric/trace 导出。"],
        ["建议先做的最小实体", "cmcc4a.service / cmcc4a.redis / cmcc4a.database / cmcc4a.mq / cmcc4a.gateway，账号实体先保留为候选。"],
        ["建议先做的最小数据集", "1个应用日志集 + 1个 APM 调用链集 + Redis/DB/MQ/Gateway 4个指标集。"],
        ["你主要要补的内容", "统一命名、主键粒度、真实生产查询入口，以及专家对象到真实 serviceName 的映射。"],
        ["注意", "metadata.name 这类技术字段仍需遵守 UnifiedModel 命名规则；本文件中多数名字都是候选值，可以直接调整。"],
    ]
    for row in readme_rows:
        ws0.append(row)
    autosize(ws0)

    add_sheet(wb, "1_MinimumModelPlan", ["Kind", "Recommended Count", "What To Build First", "Why"], [
        ["entity_set", 5, "service, redis, database, mq, gateway", "专家知识和真实观测字段都已经能支撑这些实体"],
        ["entity_set_link", 5, "business_system->service, service->redis/db/mq, gateway->service", "已经足够形成最小故障链路"],
        ["log_set", 1, "application log set", "服务日志字段比较稳定，serviceName/traceId 清楚"],
        ["trace_set", 1, "apm trace set", "traceId/spanId/parentSpanId 字段比较稳定"],
        ["metric_set", 4, "redis/database/mq/gateway", "真实标签字段已经观察到"],
        ["data_link", 8, "entity-to-log/trace/metric links", "后续查询能串起来"],
        ["external_storage", 3, "log/metric/trace source", "后续可替换成正式查询入口"],
        ["storage_link", 6, "dataset-to-storage", "把证据入口说明清楚"],
    ])
    add_sheet(wb, "2_EntitySets_ToFill", ["Priority", "metadata.domain", "metadata.name", "display_name", "对象范围", "候选主键", "候选辅助字段", "主要来源", "预填依据", "还需要补什么", "工程就绪度", "决策"], entity_rows)
    add_sheet(wb, "3_EntityLinks_ToFill", ["Priority", "metadata.name", "src entity_set", "dest entity_set", "entity_link_type", "依据", "候选 fields_mapping", "还需要补什么", "备注"], relation_rows)
    add_sheet(wb, "4_DataSets_ToFill", ["Priority", "kind", "metadata.domain", "metadata.name", "用途", "核心字段/标签", "关键配置", "当前样本来源", "已观测到的 serviceName", "还需要补什么"], dataset_rows)
    add_sheet(wb, "5_DataLinks_ToFill", ["src entity_set", "dest dataset", "data_link_type", "候选 fields_mapping", "Priority", "状态", "备注"], datalink_rows)
    add_sheet(wb, "6_Storages_ToFill", ["kind", "metadata.name", "spec.type", "spec.name", "当前本地路径", "候选 properties", "还需要补什么"], storage_rows)
    add_sheet(wb, "7_StorageLinks_ToFill", ["src dataset", "dest storage", "候选 fields_mapping", "还需要补什么"], storage_link_rows)
    add_sheet(wb, "8_ObservedFieldProfile", ["Source", "Field", "Observed In Sample", "Example Value", "Suggested Use", "Candidate Target", "Confidence", "Notes"], profile_rows)
    add_sheet(wb, "9_ServiceInventory", ["Source", "Observed serviceName", "Sample Count", "Guessed Object Type", "Expert Mapping To Fill", "Notes"], service_inventory)
    add_sheet(wb, "10_ExpertRulesAndCases", ["Source Row", "Scene/Case", "Symptom", "Main Object", "Object Type/Root Cause Type", "Evidence", "Join Keys", "Rule/Conclusion", "Common Root Cause", "Action", "Notes"], rule_rows)
    add_sheet(wb, "11_OpenQuestions", ["ID", "Topic", "Question", "Why It Matters", "Owner"], open_questions)

    output_xlsx = OUTPUTS_DIR / "cmcc4a_unifiedmodel_ontology_fill_pack.xlsx"
    try:
        wb.save(output_xlsx)
    except PermissionError:
        output_xlsx = OUTPUTS_DIR / "cmcc4a_unifiedmodel_ontology_fill_pack.zh.xlsx"
        wb.save(output_xlsx)

    summary = {
        "generated_from": {
            "project_root": str(ROOT),
            "expert_workbook": str(expert_path),
            "data_dir": str(DATA_DIR),
        },
        "minimum_entity_sets": [r[2] for r in entity_rows if r[11] in ("先做", "候选")],
        "minimum_datasets": [r[3] for r in dataset_rows],
        "open_questions": [{"id": q[0], "topic": q[1], "question": q[2]} for q in open_questions],
        "top_services": {
            kind: source_stats[kind]["services"].most_common(20) for kind in source_stats
        },
    }
    output_json = OUTPUTS_DIR / "cmcc4a_unifiedmodel_ontology_fill_pack.summary.json"
    output_json.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

    return output_xlsx, output_json


if __name__ == "__main__":
    xlsx, summary = build_workbook()
    print(xlsx)
    print(summary)
