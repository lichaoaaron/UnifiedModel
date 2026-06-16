from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "outputs" / "project-management"
XLSX_PATH = OUTPUT_DIR / "MModel项目分工与排期.xlsx"
MD_PATH = OUTPUT_DIR / "MModel项目作战说明.md"


TASK_HEADERS = [
    "工作线",
    "模块",
    "子任务",
    "负责人",
    "参与人",
    "输入",
    "输出",
    "开始时间",
    "截止时间",
    "当前状态",
    "验收标准",
]

TASK_ROWS = [
    [
        "平台主干线",
        "Schema",
        "冻结中国移动一级4A一期最小entity_set清单",
        "你",
        "平台开发A, 运维专家A, 运维专家B",
        "专家知识表, 现有MModel schema, 8个故障案例",
        "一期entity_set清单v1",
        "2026-06-11",
        "2026-06-14",
        "未开始",
        "服务/接口/数据库/Redis/Pod等核心对象范围明确，团队无歧义",
    ],
    [
        "平台主干线",
        "Schema",
        "冻结一期entity_set_link与relation_type命名规范",
        "平台开发A",
        "你, 运维专家A",
        "故障路径分析方法, 现有incident-investigation样例",
        "关系命名规范v1",
        "2026-06-13",
        "2026-06-16",
        "未开始",
        "至少覆盖service调用、service依赖db/redis、service运行在workload等关键关系",
    ],
    [
        "平台主干线",
        "Schema",
        "冻结一期dataset范围与DataLink/StorageLink约束",
        "平台开发A",
        "你, 数据工具开发",
        "日志/指标/链路字段画像结果, OpenSearch现状",
        "dataset与link规范v1",
        "2026-06-13",
        "2026-06-17",
        "未开始",
        "明确哪些log/metric/trace集合进入一期，以及如何挂接对象",
    ],
    [
        "平台主干线",
        "Import",
        "整理4A模型包目录结构并形成导入规范",
        "平台开发A",
        "你",
        "现有examples结构, 一期schema清单",
        "4A模型包目录模板",
        "2026-06-15",
        "2026-06-18",
        "未开始",
        "新成员可按模板补充模型文件，无需自行设计目录",
    ],
    [
        "平台主干线",
        "Import",
        "实现4A模型包一键导入脚本与导入校验报告",
        "平台开发A",
        "平台开发B",
        "模型包目录, mmodel import能力",
        "导入脚本, 导入校验报告",
        "2026-06-16",
        "2026-06-20",
        "未开始",
        "新workspace可以一键导入，失败时定位到具体文件",
    ],
    [
        "平台主干线",
        "GraphStore",
        "定义entity写入协议v1",
        "平台开发B",
        "你, 数据工具开发",
        "examples/sample-data/entities.json, 4A对象清单",
        "entity写入协议文档与样例",
        "2026-06-15",
        "2026-06-18",
        "未开始",
        "RCA线和工具线可以稳定产出符合协议的entity数据",
    ],
    [
        "平台主干线",
        "GraphStore",
        "定义relation写入协议v1",
        "平台开发B",
        "你, 数据工具开发",
        "examples/sample-data/relations.json, 关系命名规范",
        "relation写入协议文档与样例",
        "2026-06-15",
        "2026-06-18",
        "未开始",
        "至少支持direct relation查询和路径查询的关键字段",
    ],
    [
        "平台主干线",
        "Query",
        "实现按entity type查询对象实例能力",
        "平台开发B",
        "前端开发A, RCA开发A",
        "entity写入协议, query service现状",
        ".entity查询能力v1",
        "2026-06-18",
        "2026-06-22",
        "未开始",
        "前端能按类型列出服务、数据库、Pod等实例",
    ],
    [
        "平台主干线",
        "Query",
        "实现按对象查直接关系与邻接节点能力",
        "平台开发B",
        "前端开发A, RCA开发B",
        "relation写入协议, topo查询需求",
        "对象邻接查询API",
        "2026-06-18",
        "2026-06-23",
        "未开始",
        "给定告警对象，可以查到上下游与依赖对象",
    ],
    [
        "平台主干线",
        "API Contract",
        "冻结前端/RCA共用接口字段命名和返回结构",
        "你",
        "平台开发B, RCA开发A, 前端开发A",
        "告警演示流程, query接口, RCA输出草案",
        "API contract v1",
        "2026-06-20",
        "2026-06-24",
        "未开始",
        "前后端联调过程中不再频繁改字段名",
    ],
    [
        "RCA能力线",
        "Alert In",
        "定义告警标准输入结构",
        "RCA开发A",
        "你, 平台开发B",
        "现有告警样本, 领导演示诉求",
        "alert payload v1",
        "2026-06-17",
        "2026-06-19",
        "未开始",
        "至少包含告警对象、指标、时间窗、严重度、来源",
    ],
    [
        "RCA能力线",
        "Candidate Binding",
        "实现告警对象到图中对象的绑定",
        "RCA开发A",
        "平台开发B",
        "alert payload, entity查询能力",
        "对象绑定逻辑v1",
        "2026-06-19",
        "2026-06-24",
        "未开始",
        "给定一个告警，可定位到候选service/infra对象",
    ],
    [
        "RCA能力线",
        "Path Selection",
        "实现候选路径生成：底层资源路径/下游依赖路径/本服务异常路径",
        "RCA开发B",
        "你, 平台开发B",
        "运维专家排障经验, 对象关系图",
        "候选路径生成器v1",
        "2026-06-20",
        "2026-06-27",
        "未开始",
        "每个告警至少输出3类可解释的排障路径",
    ],
    [
        "RCA能力线",
        "Evidence",
        "实现trace/log/metric证据汇聚",
        "RCA开发A",
        "RCA开发B, 数据工具开发",
        "原始观测数据样本, dataset/link规则",
        "证据摘要结构v1",
        "2026-06-22",
        "2026-06-29",
        "未开始",
        "每条候选路径都能挂上至少一种可展示证据",
    ],
    [
        "RCA能力线",
        "Ranking",
        "实现根因候选排序与解释输出",
        "RCA开发B",
        "RCA开发A, 运维专家A",
        "证据摘要, 8个故障案例",
        "root cause ranking v1",
        "2026-06-26",
        "2026-07-03",
        "未开始",
        "能给出TopN根因候选、分数和解释理由",
    ],
    [
        "RCA能力线",
        "Validation",
        "基于8个故障案例形成回放验证结果",
        "RCA开发A",
        "你, 运维专家B",
        "8个故障案例, 证据聚合与排序能力",
        "RCA验证报告v1",
        "2026-07-01",
        "2026-07-08",
        "未开始",
        "至少3个案例形成可演示闭环，其他案例有差距说明",
    ],
    [
        "前端演示线",
        "告警入口",
        "实现告警风暴入口页",
        "前端开发A",
        "前端开发B, 你",
        "演示脚本, alert payload",
        "告警入口页v1",
        "2026-06-20",
        "2026-06-26",
        "未开始",
        "领导能一眼看到告警列表、严重度和初步聚焦对象",
    ],
    [
        "前端演示线",
        "拓扑展示",
        "实现候选路径与对象拓扑展示页",
        "前端开发A",
        "平台开发B, RCA开发B",
        "对象邻接查询API, 候选路径结构",
        "路径拓扑页v1",
        "2026-06-24",
        "2026-07-01",
        "未开始",
        "可以可视化展示服务、数据库、Redis、Pod等对象和路径",
    ],
    [
        "前端演示线",
        "证据面板",
        "实现trace/log/metric证据面板",
        "前端开发B",
        "RCA开发A, RCA开发B",
        "证据摘要结构",
        "证据面板v1",
        "2026-06-26",
        "2026-07-03",
        "未开始",
        "用户能直接看到关键trace异常、关键日志和异常指标",
    ],
    [
        "前端演示线",
        "结论页",
        "实现根因结论与解释页",
        "前端开发B",
        "RCA开发B, 你",
        "root cause ranking结构, 演示文案",
        "根因结论页v1",
        "2026-07-01",
        "2026-07-05",
        "未开始",
        "能清楚展示Top1/Top2候选、置信度和依据",
    ],
    [
        "前端演示线",
        "Dashboard联动",
        "实现OpenSearch Dashboard嵌入和跳转联动",
        "前端开发A",
        "前端开发B",
        "现有dashboard嵌入能力, query参数约定",
        "Dashboard联动v1",
        "2026-07-02",
        "2026-07-08",
        "未开始",
        "结论页可一键跳到对应日志/指标/链路明细",
    ],
    [
        "数据建模与工具线",
        "字段画像",
        "扫描log/metric/trace字段分布并输出画像",
        "数据工具开发",
        "你",
        "本地OpenSearch导出数据",
        "字段画像报告v1",
        "2026-06-11",
        "2026-06-16",
        "未开始",
        "能看出稳定键、候选实体字段、可关联字段",
    ],
    [
        "数据建模与工具线",
        "候选建模",
        "生成候选entity_set/dataset/datalink建议清单",
        "数据工具开发",
        "你, 平台开发A",
        "字段画像报告, 现有专家表",
        "候选建模建议v1",
        "2026-06-16",
        "2026-06-20",
        "未开始",
        "至少能覆盖服务、接口、数据库、Redis、Pod等一期对象",
    ],
    [
        "数据建模与工具线",
        "表转模型",
        "实现专家Excel到YAML/JSON模型包转换脚本",
        "数据工具开发",
        "平台开发A",
        "专家填表, 模型包目录规范",
        "转换脚本v1",
        "2026-06-20",
        "2026-06-27",
        "未开始",
        "专家更新表后可以半自动生成模型包草案",
    ],
    [
        "数据建模与工具线",
        "故障样本",
        "实现8个故障案例回放样本整理与复用脚本",
        "数据工具开发",
        "RCA开发A, 你",
        "故障案例, 原始观测数据, 现有fault samples脚本",
        "回放样本包v1",
        "2026-06-24",
        "2026-07-02",
        "未开始",
        "可以稳定回放至少3个演示案例",
    ],
]


RACI_HEADERS = [
    "事项",
    "你（项目负责人）",
    "平台主干线",
    "RCA能力线",
    "前端演示线",
    "数据建模与工具线",
    "运维专家A",
    "运维专家B",
]

RACI_ROWS = [
    ["8月底目标与边界定义", "A/R", "C", "C", "C", "C", "I", "I"],
    ["一期4A最小schema范围冻结", "A/R", "R", "C", "I", "C", "C", "C"],
    ["关系命名规范冻结", "A", "R", "C", "I", "I", "C", "C"],
    ["模型包导入规范", "A", "R", "I", "I", "C", "I", "I"],
    ["entity/relation写入协议", "A", "R", "C", "I", "C", "I", "I"],
    ["统一query/API contract", "A", "R", "C", "C", "I", "I", "I"],
    ["告警输入结构定义", "A", "C", "R", "C", "I", "I", "I"],
    ["候选路径生成策略", "A", "C", "R", "I", "I", "C", "C"],
    ["trace/log/metric证据策略", "A", "C", "R", "I", "C", "C", "C"],
    ["告警风暴演示页面", "A", "I", "C", "R", "I", "I", "I"],
    ["拓扑和证据可视化", "A", "C", "C", "R", "I", "I", "I"],
    ["OpenSearch Dashboard联动", "A", "C", "I", "R", "I", "I", "I"],
    ["字段画像与候选建模建议", "A", "C", "I", "I", "R", "I", "I"],
    ["专家表转模型包脚本", "A", "C", "I", "I", "R", "I", "I"],
    ["8个故障案例验收", "A/R", "C", "R", "C", "C", "C", "C"],
    ["演示分支收口", "A/R", "C", "C", "C", "I", "I", "I"],
]


MILESTONE_HEADERS = ["里程碑", "目标", "负责人", "截止时间", "验收方式", "风险"]

MILESTONE_ROWS = [
    [
        "M1 范围冻结",
        "冻结一期对象范围、关系范围、dataset范围和统一命名",
        "你",
        "2026-06-17",
        "评审会通过并发布v1文档",
        "范围失控，专家补充过多导致持续返工",
    ],
    [
        "M2 模型导入闭环",
        "4A模型包可一键导入workspace并通过校验",
        "平台开发A",
        "2026-06-20",
        "新workspace导入成功且输出校验报告",
        "schema不稳定导致导入脚本频繁改动",
    ],
    [
        "M3 图写入闭环",
        "entity/relation样本可写入并可查询",
        "平台开发B",
        "2026-06-23",
        "可通过.entity/.topo查到4A核心对象与关系",
        "协议不稳定导致RCA/前端无法并行",
    ],
    [
        "M4 RCA最小闭环",
        "告警输入到候选路径和根因候选可跑通",
        "RCA开发A",
        "2026-07-03",
        "至少1个故障案例自动生成候选路径和TopN根因",
        "证据不足或路径解释不清",
    ],
    [
        "M5 演示界面闭环",
        "告警、路径、证据、结论四个页面串起来",
        "前端开发A",
        "2026-07-08",
        "团队内部可演示一遍完整链路",
        "接口字段变化影响联调",
    ],
    [
        "M6 三案例稳定回放",
        "至少3个案例可稳定回放展示",
        "你",
        "2026-07-15",
        "连续3次演示结果一致",
        "样本不稳定或环境差异导致回放漂移",
    ],
    [
        "M7 演示分支收口",
        "演示版功能冻结，只修bug不加大功能",
        "你",
        "2026-08-15",
        "release分支冻结并通过彩排",
        "后期还在改大功能导致质量失控",
    ],
]


RISK_HEADERS = ["问题/依赖", "影响范围", "当前负责人", "需谁决策", "期望解决时间"]

RISK_ROWS = [
    ["专家表尚未填完", "schema冻结、关系定义、RCA规则", "你", "你", "2026-06-14"],
    ["原始观测数据字段命名不统一", "DataLink设计、证据聚合、对象绑定", "数据工具开发", "你", "2026-06-16"],
    ["告警对象与图中实体映射不稳定", "RCA入口、前端演示可信度", "RCA开发A", "你", "2026-06-20"],
    ["OpenSearch Dashboard联动参数未统一", "前端跳转体验", "前端开发A", "你", "2026-07-01"],
    ["故障样例缺少明确真值", "RCA验收标准", "你", "运维专家A/运维专家B", "2026-06-24"],
    ["多人并行开发导致接口频繁变动", "联调效率、演示稳定性", "平台开发B", "你", "持续跟踪"],
]


MD_CONTENT = """# MModel项目作战说明

## 1. 项目目标

8月底前，围绕中国移动一级4A场景，完成一个可演示、可解释、可复现的MModel最小闭环：

1. 输入一个真实告警或故障样例。
2. 系统自动定位相关对象。
3. 系统自动给出最可能的多条排障路径。
4. 系统自动汇聚trace/log/metric证据。
5. 系统给出TopN根因候选和解释说明。
6. 页面支持跳转到OpenSearch Dashboard继续深挖。

核心不是“全自动替代专家”，而是“自动收敛路径、自动聚合证据、明显提升排障效率”。

## 2. 8月底演示边界

本阶段必须做：

- 中国移动一级4A一期最小对象范围。
- 至少3个故障案例可稳定回放。
- 告警 -> 路径 -> 证据 -> 根因候选 的完整界面链路。
- 根因候选必须有解释，而不是只给一个结论。

本阶段明确不做：

- 全量4A本体一次性建完。
- 所有告警类型全面覆盖。
- 完全替代运维专家人工判断。
- 大规模在线实时生产化承诺。

## 3. 四条工作线职责边界

### 3.1 平台主干线

负责“系统认什么对象、这些对象如何导入、如何写入图、如何被统一查询出来”。

包括：

- schema
- import
- graphstore
- query
- API contract

这条线是项目地基。其他线原则上不直接修改核心契约。

### 3.2 RCA能力线

负责“告警进来以后，如何自动往下钻，如何选路径，如何汇聚证据，如何给出根因候选”。

包括：

- 告警标准化
- 对象绑定
- 候选路径生成
- trace/log/metric证据汇聚
- 根因候选排序
- 解释输出

### 3.3 前端演示线

负责“把复杂的排障过程讲清楚，让领导一眼看到价值”。

包括：

- 告警入口页
- 路径拓扑页
- 证据面板
- 根因结论页
- Dashboard嵌入与跳转

### 3.4 数据建模与工具线

负责“把专家知识和真实观测数据翻译成平台可消费的结构化产物”。

包括：

- 字段画像
- 候选entity_set/dataset/datalink建议
- 专家Excel到模型包转换
- 故障案例回放样本生成

## 4. 分支与合并规则

建议采用四层分支：

1. `main`
2. `integration/platform`
3. `integration/rca`
4. `integration/ui`
5. `release/demo-aug`

规则如下：

- 日常开发先进入各自 `integration/*`。
- 每周固定两次从 `integration/*` 合并到 `main`。
- `main` 必须保持可启动、可导入、可演示基础能力不坏。
- 演示前两周，`release/demo-aug` 只收敛不扩需求。
- 大的schema/API改动必须先过负责人评审，再进入开发。

## 5. 会议与节奏

建议固定以下节奏：

1. 每天15分钟站会  
同步昨天完成、今天计划、当前阻塞。

2. 每周一次架构/契约评审  
只讨论边界、契约、风险，不卷实现细节。

3. 每周一次演示彩排  
哪怕功能不全，也要持续串链路，尽早暴露断点。

4. 每周一次对上汇报  
只汇报里程碑、风险、需要支持的决策点。

## 6. 项目负责人职责

你必须亲自承担的事项：

- 定义8月底目标、边界和成功标准。
- 冻结一期4A最小对象范围。
- 冻结命名规范和跨组契约。
- 拍板哪些功能可以进演示分支。
- 抓3个核心故障案例的闭环验收。
- 统一对领导和对专家的话术。

你不应下放的不是“所有工作”，而是“范围、契约、优先级和收口权”。

## 7. 当前最优先的五件事

1. 冻结一期对象与关系范围。
2. 冻结entity/relation/API contract。
3. 从8个故障案例里选出3个演示金样本。
4. 让字段画像和候选建模建议先跑出来。
5. 尽快形成第一版告警 -> 路径 -> 证据 -> 根因候选闭环。
"""


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def auto_width(ws, widths):
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def build_workbook():
    wb = Workbook()

    ws = wb.active
    ws.title = "任务分工表"
    ws.append(TASK_HEADERS)
    for row in TASK_ROWS:
        ws.append(row)
    style_sheet(ws)
    auto_width(ws, [14, 14, 30, 14, 24, 28, 24, 12, 12, 10, 38])

    ws = wb.create_sheet("责任矩阵")
    ws.append(RACI_HEADERS)
    for row in RACI_ROWS:
        ws.append(row)
    style_sheet(ws)
    auto_width(ws, [28, 18, 14, 14, 14, 18, 12, 12])

    ws = wb.create_sheet("里程碑")
    ws.append(MILESTONE_HEADERS)
    for row in MILESTONE_ROWS:
        ws.append(row)
    style_sheet(ws)
    auto_width(ws, [16, 34, 14, 12, 24, 28])

    ws = wb.create_sheet("问题与依赖")
    ws.append(RISK_HEADERS)
    for row in RISK_ROWS:
        ws.append(row)
    style_sheet(ws)
    auto_width(ws, [28, 26, 14, 14, 14])

    return wb


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    build_workbook().save(XLSX_PATH)
    MD_PATH.write_text(MD_CONTENT, encoding="utf-8")
    print(f"Generated: {XLSX_PATH}")
    print(f"Generated: {MD_PATH}")


if __name__ == "__main__":
    main()
