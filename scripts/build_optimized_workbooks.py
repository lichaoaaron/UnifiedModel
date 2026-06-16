#!/usr/bin/env python
# -*- coding: utf-8 -*-

import pathlib
from copy import copy
import shutil

import openpyxl


ROOT = pathlib.Path(r"C:\Users\chaoJ\Desktop\MModel")
OUTPUTS = ROOT / "outputs"


def clone_workbook(src_path: pathlib.Path, dst_path: pathlib.Path):
    shutil.copy2(src_path, dst_path)


def header_map(ws):
    return {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}


def find_row_by_value(ws, col_name, target):
    hm = header_map(ws)
    col = hm[col_name]
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, col).value == target:
            return r
    raise KeyError(f"Row not found in {ws.title}: {col_name}={target}")


def setv(ws, row, col_name, value):
    hm = header_map(ws)
    ws.cell(row, hm[col_name], value)


def optimize_entityset_sheet(ws):
    row = find_row_by_value(ws, "metadata.name", "cmcc4a.service")
    setv(ws, row, "候选辅助字段", "resource.attributes.service@name, resource.attributes.host@ip, traceId")
    setv(ws, row, "预填依据", "专家表第6-18行已经覆盖；真实观测数据里也已有大量 serviceName。当前未看到稳定的 routeId；appId 虽在部分日志/SQL中出现，但更像应用对象字段，不建议默认作为 service 辅助字段。")
    setv(ws, row, "还需要补什么", "确认统一命名、中文别名，以及 service 的辅助字段是否只保留服务侧稳定字段。")

    row = find_row_by_value(ws, "metadata.name", "cmcc4a.redis")
    setv(ws, row, "候选辅助字段", "resource.attributes.service@instance@id, metric.attributes.endpoint, serviceName")
    setv(ws, row, "预填依据", "专家表第9/12行覆盖；metric 里已看到 redis 实例 ID；trace 里有 Redis 调用。当前未看到稳定的 keyPrefix 观测字段。")

    row = find_row_by_value(ws, "metadata.name", "cmcc4a.database")
    setv(ws, row, "候选主键", "instanceId")
    setv(ws, row, "候选辅助字段", "span.attributes.db@instance, span.attributes.db@statement, resource.attributes.service@instance@id")
    setv(ws, row, "还需要补什么", "确认数据库是否先按 instanceId 建模；如后续能稳定补出 dbName/clusterId，再迭代增强。")

    row = find_row_by_value(ws, "metadata.name", "cmcc4a.mq")
    setv(ws, row, "候选主键", "topic or consumerGroup")
    setv(ws, row, "候选辅助字段", "metric.attributes.topic, metric.attributes.group, metric.attributes.clientId, metric.attributes.cmd")
    setv(ws, row, "预填依据", "专家表第10/12行覆盖；metric 里已看到 topic/group/clientId/cmd。当前未看到稳定的 jobName 字段。")

    row = find_row_by_value(ws, "metadata.name", "cmcc4a.gateway")
    setv(ws, row, "候选主键", "resource.attributes.service@instance@id or metric.attributes.name")
    setv(ws, row, "候选辅助字段", "metric.attributes.endpoint, metric.attributes.name, metric.attributes.status, metric.attributes.role, metric.attributes.url, metric.attributes.backend")
    setv(ws, row, "还需要补什么", "建议一期先按网关实例/上游目标建模，不默认引入 routeId；如果后续确认真实路由字段，再细化。")

    row = find_row_by_value(ws, "metadata.name", "cmcc4a.business_system")
    setv(ws, row, "候选主键", "appId（如确认它代表业务应用） or 专家命名的系统名称")
    setv(ws, row, "候选辅助字段", "appCode, tenantId")


def optimize_entitylink_sheet(ws):
    row = find_row_by_value(ws, "metadata.name", "cmcc4a.business_system_calls_cmcc4a.service")
    setv(ws, row, "候选 fields_mapping", "appId/appCode -> serviceName（待专家确认 appId/appCode 是否可代表业务应用）")

    row = find_row_by_value(ws, "metadata.name", "cmcc4a.gateway_serves_cmcc4a.service")
    setv(ws, row, "候选 fields_mapping", "metric.attributes.backend or metric.attributes.name -> serviceName（先去掉 routeId）")

    row = find_row_by_value(ws, "metadata.name", "cmcc4a.account_access_cmcc4a.business_system")
    setv(ws, row, "候选 fields_mapping", "userId/accountId -> appId/appCode（仅在账号实体正式入模时再做）")


def optimize_dataset_sheet(ws):
    row = find_row_by_value(ws, "metadata.name", "cmcc4a.metric.gateway")
    setv(ws, row, "还需要补什么", "确认一期是否只按网关实例、upstream name、backend、url 建模；routeId 暂不默认纳入。")


def optimize_datalink_sheet(ws):
    row = find_row_by_value(ws, "dest dataset", "cmcc4a.metric.gateway")
    if ws.cell(row, header_map(ws)["src entity_set"]).value == "cmcc4a.gateway":
        setv(ws, row, "候选 fields_mapping", "metric.attributes.name/url/backend -> gateway instance or upstream target（先去掉 routeId）")


def optimize_openquestions_sheet(ws):
    row = find_row_by_value(ws, "ID", "Q4")
    setv(ws, row, "Question", "是否一期先按网关实例/上游目标建模，不默认引入 routeId；若后续拿到稳定路由字段，再细化为 route 子实体。")


def optimize_full_confirmation_book(path: pathlib.Path):
    wb = openpyxl.load_workbook(str(path))
    optimize_entityset_sheet(wb["2_EntitySets_ToFill"])
    optimize_entitylink_sheet(wb["3_EntityLinks_ToFill"])
    optimize_dataset_sheet(wb["4_DataSets_ToFill"])
    optimize_datalink_sheet(wb["5_DataLinks_ToFill"])
    optimize_openquestions_sheet(wb["11_OpenQuestions"])
    wb.save(path)


def optimize_minimal_book(path: pathlib.Path):
    wb = openpyxl.load_workbook(str(path))

    ws = wb["1_核心实体确认"]
    hm = header_map(ws)
    row = find_row_by_value(ws, "metadata.name", "cmcc4a.service")
    setv(ws, row, "候选辅助字段", "resource.attributes.service@name, resource.attributes.host@ip, traceId")
    setv(ws, row, "预填依据", "专家表第6-18行已经覆盖；真实观测数据里也已有大量 serviceName。当前未看到稳定的 routeId；appId 虽在部分日志/SQL中出现，但更像应用对象字段。")
    setv(ws, row, "还需要补什么", "确认 service 的辅助字段是否只保留服务侧稳定字段。")

    row = find_row_by_value(ws, "metadata.name", "cmcc4a.redis")
    setv(ws, row, "候选辅助字段", "resource.attributes.service@instance@id, metric.attributes.endpoint, serviceName")

    row = find_row_by_value(ws, "metadata.name", "cmcc4a.database")
    setv(ws, row, "候选主键", "instanceId")
    setv(ws, row, "候选辅助字段", "span.attributes.db@instance, span.attributes.db@statement, resource.attributes.service@instance@id")

    row = find_row_by_value(ws, "metadata.name", "cmcc4a.mq")
    setv(ws, row, "候选主键", "topic or consumerGroup")
    setv(ws, row, "候选辅助字段", "metric.attributes.topic, metric.attributes.group, metric.attributes.clientId, metric.attributes.cmd")

    row = find_row_by_value(ws, "metadata.name", "cmcc4a.gateway")
    setv(ws, row, "候选主键", "resource.attributes.service@instance@id or metric.attributes.name")
    setv(ws, row, "候选辅助字段", "metric.attributes.endpoint, metric.attributes.name, metric.attributes.status, metric.attributes.role, metric.attributes.url, metric.attributes.backend")

    row = find_row_by_value(ws, "metadata.name", "cmcc4a.account")
    setv(ws, row, "专家备注/待确认", "如专家认为账号暂时不宜入模，可直接写“建议一期不入模”。")

    ws = wb["2_核心关系确认"]
    row = find_row_by_value(ws, "metadata.name", "cmcc4a.business_system_calls_cmcc4a.service")
    setv(ws, row, "候选 fields_mapping", "appId/appCode -> serviceName（待专家确认 appId/appCode 是否可代表业务应用）")
    row = find_row_by_value(ws, "metadata.name", "cmcc4a.gateway_serves_cmcc4a.service")
    setv(ws, row, "候选 fields_mapping", "metric.attributes.backend or metric.attributes.name -> serviceName（先去掉 routeId）")

    ws = wb["4_关键问题结论"]
    row = find_row_by_value(ws, "ID", "Q4")
    setv(ws, row, "Question", "是否一期先按网关实例/上游目标建模，不默认引入 routeId；若后续拿到稳定路由字段，再细化为 route 子实体。")

    wb.save(path)


def main():
    full_src = OUTPUTS / "cmcc4a_expert_confirmation_pack.xlsx"
    full_dst = OUTPUTS / "cmcc4a_expert_confirmation_pack.optimized.xlsx"
    clone_workbook(full_src, full_dst)
    optimize_full_confirmation_book(full_dst)

    minimal_src = OUTPUTS / "cmcc4a_expert_minimal_confirmation_pack.xlsx"
    minimal_dst = OUTPUTS / "cmcc4a_expert_minimal_confirmation_pack.optimized.xlsx"
    clone_workbook(minimal_src, minimal_dst)
    optimize_minimal_book(minimal_dst)

    print(full_dst)
    print(minimal_dst)


if __name__ == "__main__":
    main()
