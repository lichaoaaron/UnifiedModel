#!/usr/bin/env python
# -*- coding: utf-8 -*-

import pathlib
from copy import copy

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = pathlib.Path(r"C:\Users\chaoJ\Desktop\UnifiedModel")
OUTPUTS = ROOT / "outputs"
SOURCE = OUTPUTS / "cmcc4a_expert_confirmation_pack.xlsx"
TARGET = OUTPUTS / "cmcc4a_expert_minimal_confirmation_pack.xlsx"

BLUE = PatternFill("solid", fgColor="D9EAF7")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
GREEN = PatternFill("solid", fgColor="E2F0D9")
RED = PatternFill("solid", fgColor="F4CCCC")
THIN = Side(style="thin", color="BFBFBF")


def autosize(ws):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 42)


def add_dropdown(ws, sqref, options):
    formula = '"' + ",".join(options) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.prompt = "请从下拉框中选择"
    dv.error = "请选择下拉框中的值"
    ws.add_data_validation(dv)
    dv.add(sqref)


def copy_header_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell._style = copy(src_cell._style)
    dst_cell.font = Font(bold=True)
    dst_cell.alignment = Alignment(wrap_text=True, vertical="center")
    dst_cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def build_guide_sheet(wb):
    ws = wb.active
    ws.title = "0_请专家确认_最小版"
    ws.append(["项目", "填写说明"])
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)
    ws["A1"].fill = BLUE
    ws["B1"].fill = BLUE
    rows = [
        ("目标", "这份表只保留最关键确认项，目的是尽快拿到可以启动 4A 最小本体构建的专家结论。"),
        ("请优先确认", "1. 核心实体是否对。2. 核心关系方向是否对。3. serviceName 对照是否对。4. 关键开放问题是否能给出结论。"),
        ("专家怎么填", "如果内容正确，选“直接确认/已确认”。如果不对，直接改原值，并在右侧写修改说明。暂时不能定的，选“暂时不能确认”，并说明还缺什么。"),
        ("工作量", "这版就是为了把专家工作量压到最小，不求一次补全全部细节。"),
        ("回传标准", "关键行都填了“专家处理结果”；有修改的地方写了修改说明；拿不准的地方写了备注。"),
    ]
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    autosize(ws)


def build_sheet_from_source(dst_ws, src_ws, keep_columns, keep_rows, focus_text):
    dst_headers = [src_ws.cell(row=1, column=c).value for c in keep_columns]
    for col_idx, header in enumerate(dst_headers, start=1):
        cell = dst_ws.cell(row=1, column=col_idx, value=header)
        copy_header_style(src_ws.cell(row=1, column=keep_columns[col_idx - 1]), cell)
        if header in {"display_name", "对象范围", "候选主键", "候选 fields_mapping", "Question", "Observed serviceName", "Expert Mapping To Fill"}:
            cell.fill = YELLOW

    new_base = len(keep_columns) + 1
    extra_headers = [
        "专家操作建议",
        "需重点确认什么",
        "专家处理结果",
        "如需修改，请写修改后的值",
        "如需补充，请写补充内容",
        "修改/补充说明",
        "专家备注/待确认",
    ]
    for i, header in enumerate(extra_headers, start=new_base):
        cell = dst_ws.cell(row=1, column=i, value=header)
        cell.font = Font(bold=True)
        cell.fill = GREEN
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    action_col = get_column_letter(new_base)
    result_col = get_column_letter(new_base + 2)

    out_row = 2
    for src_row in keep_rows:
        for dst_col_idx, src_col_idx in enumerate(keep_columns, start=1):
            src_cell = src_ws.cell(row=src_row, column=src_col_idx)
            dst_cell = dst_ws.cell(row=out_row, column=dst_col_idx, value=src_cell.value)
            if src_cell.has_style:
                dst_cell._style = copy(src_cell._style)
            dst_cell.alignment = Alignment(wrap_text=True, vertical="top")
            dst_cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

        dst_ws.cell(row=out_row, column=new_base, value="直接确认")
        dst_ws.cell(row=out_row, column=new_base + 1, value=focus_text(src_ws, src_row))
        dst_ws.cell(row=out_row, column=new_base + 2, value="待确认")

        for c in range(new_base, new_base + 7):
            cell = dst_ws.cell(row=out_row, column=c)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            if c >= new_base + 2:
                cell.fill = YELLOW
        dst_ws.cell(row=out_row, column=new_base + 1).fill = RED
        out_row += 1

    if out_row > 2:
        add_dropdown(dst_ws, f"{action_col}2:{action_col}{out_row-1}", ["直接确认", "修改原值", "补充信息", "暂时不能确认"])
        add_dropdown(dst_ws, f"{result_col}2:{result_col}{out_row-1}", ["已确认", "已修改", "已补充", "需工程再确认", "需更多信息"])

    dst_ws.freeze_panes = "A2"
    autosize(dst_ws)


def find_rows_by_values(ws, col_idx, values):
    rows = []
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, col_idx).value in values:
            rows.append(r)
    return rows


def main():
    if not SOURCE.exists():
        raise FileNotFoundError(f"Source workbook not found: {SOURCE}")

    src_wb = openpyxl.load_workbook(str(SOURCE), data_only=False)
    wb = Workbook()
    build_guide_sheet(wb)

    # 1. core entities
    src_entities = src_wb["2_EntitySets_ToFill"]
    keep_entity_rows = find_rows_by_values(src_entities, 3, [
        "cmcc4a.service",
        "cmcc4a.redis",
        "cmcc4a.database",
        "cmcc4a.mq",
        "cmcc4a.gateway",
        "cmcc4a.account",
    ])
    ws = wb.create_sheet("1_核心实体确认")
    build_sheet_from_source(
        ws,
        src_entities,
        keep_columns=[1, 3, 4, 5, 6, 7, 9, 10, 12],
        keep_rows=keep_entity_rows,
        focus_text=lambda _ws, _r: "请重点确认：对象中文名、对象范围、候选主键是否正确。账号实体如果不建议现在入模，请直接说明。",
    )

    # 2. core relations
    src_links = src_wb["3_EntityLinks_ToFill"]
    keep_link_rows = find_rows_by_values(src_links, 2, [
        "cmcc4a.business_system_calls_cmcc4a.service",
        "cmcc4a.service_use_cmcc4a.redis",
        "cmcc4a.service_use_cmcc4a.database",
        "cmcc4a.service_sends_to_cmcc4a.mq",
        "cmcc4a.gateway_serves_cmcc4a.service",
    ])
    ws = wb.create_sheet("2_核心关系确认")
    build_sheet_from_source(
        ws,
        src_links,
        keep_columns=[2, 3, 4, 5, 6, 7, 8, 9],
        keep_rows=keep_link_rows,
        focus_text=lambda _ws, _r: "请重点确认：这条关系是否成立，方向是否正确，关联字段是否合理。",
    )

    # 3. serviceName mapping
    src_services = src_wb["9_ServiceInventory"]
    keep_service_rows = []
    seen = 0
    for r in range(2, src_services.max_row + 1):
        source = src_services.cell(r, 1).value
        if source != "log":
            continue
        keep_service_rows.append(r)
        seen += 1
        if seen >= 20:
            break
    ws = wb.create_sheet("3_serviceName对照确认")
    build_sheet_from_source(
        ws,
        src_services,
        keep_columns=[1, 2, 3, 4, 5],
        keep_rows=keep_service_rows,
        focus_text=lambda _ws, _r: "请把这个 serviceName 对应到专家口径中的对象名；如果看不出来，请写备注。",
    )

    # 4. key open questions
    src_q = src_wb["11_OpenQuestions"]
    keep_q_rows = list(range(2, src_q.max_row + 1))
    ws = wb.create_sheet("4_关键问题结论")
    build_sheet_from_source(
        ws,
        src_q,
        keep_columns=[1, 2, 3, 4, 5],
        keep_rows=keep_q_rows,
        focus_text=lambda _ws, _r: "请尽量直接给结论；如果暂时不能定，请写清楚还缺什么信息。",
    )

    wb.save(TARGET)
    print(TARGET)


if __name__ == "__main__":
    main()
