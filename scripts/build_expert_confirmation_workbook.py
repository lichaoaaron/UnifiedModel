#!/usr/bin/env python
# -*- coding: utf-8 -*-

import pathlib
from copy import copy

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = pathlib.Path(r"C:\Users\chaoJ\Desktop\MModel")
OUTPUTS = ROOT / "outputs"
SOURCE = OUTPUTS / "cmcc4a_mmodel_ontology_fill_pack.zh.xlsx"
TARGET = OUTPUTS / "cmcc4a_expert_confirmation_pack.xlsx"

YELLOW = PatternFill("solid", fgColor="FFF2CC")
BLUE = PatternFill("solid", fgColor="D9EAF7")
GREEN = PatternFill("solid", fgColor="E2F0D9")
LIGHT_GREEN = PatternFill("solid", fgColor="EAF4E6")
STRONG_RED = PatternFill("solid", fgColor="F4CCCC")
THIN = Side(style="thin", color="BFBFBF")


def zh(text: str) -> str:
    return text.encode("utf-8").decode("utf-8")


def autosize(ws):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 42)


def copy_sheet_values(src_ws, dest_ws):
    for row in src_ws.iter_rows():
        for cell in row:
            new_cell = dest_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell._style = copy(cell._style)
            if cell.font:
                new_cell.font = copy(cell.font)
            if cell.fill:
                new_cell.fill = copy(cell.fill)
            if cell.border:
                new_cell.border = copy(cell.border)
            if cell.alignment:
                new_cell.alignment = copy(cell.alignment)
            if cell.number_format:
                new_cell.number_format = cell.number_format
    for key, dim in src_ws.column_dimensions.items():
        dest_ws.column_dimensions[key].width = dim.width
    for key, dim in src_ws.row_dimensions.items():
        dest_ws.row_dimensions[key].height = dim.height
    if src_ws.freeze_panes:
        dest_ws.freeze_panes = src_ws.freeze_panes


def add_dropdown(ws, sqref, options):
    formula = '"' + ",".join(options) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.prompt = zh("\u8bf7\u4ece\u4e0b\u62c9\u6846\u4e2d\u9009\u62e9")
    dv.error = zh("\u8bf7\u9009\u62e9\u4e0b\u62c9\u6846\u4e2d\u7684\u503c")
    ws.add_data_validation(dv)
    dv.add(sqref)


def mark_uncertain_text(value):
    if value is None:
        return False
    text = str(value)
    flags = [
        zh("\u786e\u8ba4"),
        zh("\u5019\u9009"),
        zh("\u5f85"),
        zh("\u9700\u8981"),
        zh("\u662f\u5426"),
        " or ",
        "mapping",
        "topic",
        "consumerGroup",
        "routeId",
        "clusterId",
        "instanceId",
    ]
    return any(flag in text for flag in flags)


def highlight_columns(ws, headers):
    header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    for header in headers:
        col = header_map.get(header)
        if not col:
            continue
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = YELLOW
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def highlight_uncertain_cells(ws, headers):
    header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    for header in headers:
        col = header_map.get(header)
        if not col:
            continue
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            if mark_uncertain_text(cell.value):
                cell.fill = STRONG_RED
                cell.font = Font(color="9C0006", bold=True)
                cell.alignment = Alignment(wrap_text=True, vertical="top")


def add_confirmation_columns(ws, focus_text_getter):
    base_col = ws.max_column + 1
    headers = [
        zh("\u4e13\u5bb6\u64cd\u4f5c\u5efa\u8bae"),
        zh("\u9700\u91cd\u70b9\u786e\u8ba4\u4ec0\u4e48"),
        zh("\u4e13\u5bb6\u5904\u7406\u7ed3\u679c"),
        zh("\u5982\u9700\u4fee\u6539\uff0c\u8bf7\u5199\u4fee\u6539\u540e\u7684\u503c"),
        zh("\u5982\u9700\u8865\u5145\uff0c\u8bf7\u5199\u8865\u5145\u5185\u5bb9"),
        zh("\u4fee\u6539/\u8865\u5145\u8bf4\u660e"),
        zh("\u4e13\u5bb6\u5907\u6ce8/\u5f85\u786e\u8ba4"),
    ]
    for offset, header in enumerate(headers):
        cell = ws.cell(row=1, column=base_col + offset, value=header)
        cell.font = Font(bold=True)
        cell.fill = GREEN
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    action_col = base_col
    focus_col = base_col + 1
    result_col = base_col + 2
    revised_col = base_col + 3
    supplement_col = base_col + 4
    explain_col = base_col + 5
    remark_col = base_col + 6

    add_dropdown(
        ws,
        f"{get_column_letter(action_col)}2:{get_column_letter(action_col)}{ws.max_row}",
        [
            zh("\u76f4\u63a5\u786e\u8ba4"),
            zh("\u4fee\u6539\u539f\u503c"),
            zh("\u8865\u5145\u4fe1\u606f"),
            zh("\u6682\u65f6\u4e0d\u80fd\u786e\u8ba4"),
        ],
    )
    add_dropdown(
        ws,
        f"{get_column_letter(result_col)}2:{get_column_letter(result_col)}{ws.max_row}",
        [
            zh("\u5df2\u786e\u8ba4"),
            zh("\u5df2\u4fee\u6539"),
            zh("\u5df2\u8865\u5145"),
            zh("\u9700\u5de5\u7a0b\u518d\u786e\u8ba4"),
            zh("\u9700\u66f4\u591a\u4fe1\u606f"),
        ],
    )

    for row in range(2, ws.max_row + 1):
        has_content = any(ws.cell(row=row, column=c).value not in (None, "") for c in range(1, base_col))
        if not has_content:
            continue
        ws.cell(row=row, column=action_col, value=zh("\u76f4\u63a5\u786e\u8ba4"))
        ws.cell(row=row, column=focus_col, value=focus_text_getter(row))
        ws.cell(row=row, column=result_col, value=zh("\u5f85\u786e\u8ba4"))
        for c in range(base_col, base_col + 7):
            cell = ws.cell(row=row, column=c)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        ws.cell(row=row, column=focus_col).fill = LIGHT_GREEN
        ws.cell(row=row, column=result_col).fill = YELLOW
        ws.cell(row=row, column=revised_col).fill = YELLOW
        ws.cell(row=row, column=supplement_col).fill = YELLOW
        ws.cell(row=row, column=explain_col).fill = YELLOW
        ws.cell(row=row, column=remark_col).fill = YELLOW


def build_guide_sheet(out_wb):
    ws = out_wb.create_sheet(zh("0_\u8bf7\u4e13\u5bb6\u786e\u8ba4"))
    ws.append([zh("\u9879\u76ee"), zh("\u586b\u5199\u8bf4\u660e")])
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)
    ws["A1"].fill = BLUE
    ws["B1"].fill = BLUE

    rows = [
        (
            zh("\u8bf4\u660e"),
            zh("\u8fd9\u4efd\u8868\u662f\u7ed9\u4e13\u5bb6\u76f4\u63a5\u786e\u8ba4\u548c\u4fee\u6539\u7528\u7684\u3002\u9ec4\u8272\u5217\u8868\u793a\u5efa\u8bae\u91cd\u70b9\u770b\uff1b\u7ea2\u8272\u5355\u5143\u683c\u8868\u793a\u5927\u6982\u7387\u9700\u8981\u91cd\u70b9\u786e\u8ba4\u6216\u8c03\u6574\uff1b\u7eff\u8272\u533a\u57df\u662f\u8bf7\u4e13\u5bb6\u586b\u5199\u7684\u786e\u8ba4\u7ed3\u679c\u533a\u3002"),
        ),
        (
            zh("\u4e13\u5bb6\u600e\u4e48\u586b"),
            zh("1. \u5148\u770b\u7ea2\u8272\u548c\u9ec4\u8272\u5185\u5bb9\u3002 2. \u5728\u2018\u4e13\u5bb6\u64cd\u4f5c\u5efa\u8bae\u2019\u4e2d\u4ece\u4e0b\u62c9\u6846\u9009\u62e9\u3002 3. \u5982\u679c\u9700\u8981\u4fee\u6539\uff0c\u8bf7\u4f18\u5148\u76f4\u63a5\u6539\u539f\u5355\u5143\u683c\uff0c\u540c\u65f6\u5728\u53f3\u4fa7\u5199\u2018\u4fee\u6539\u540e\u7684\u503c\u2019\u548c\u2018\u4fee\u6539/\u8865\u5145\u8bf4\u660e\u2019\u3002 4. \u6682\u65f6\u62ff\u4e0d\u51c6\u7684\uff0c\u8bf7\u9009\u2018\u6682\u65f6\u4e0d\u80fd\u786e\u8ba4\u2019\uff0c\u5e76\u5728\u5907\u6ce8\u91cc\u5199\u6e05\u695a\u8fd8\u7f3a\u4ec0\u4e48\u3002"),
        ),
        (zh("\u4f18\u5148\u786e\u8ba4 1"), "2_EntitySets_ToFill\uff1a\u786e\u8ba4\u5bf9\u8c61\u540d\uff0c\u5bf9\u8c61\u8303\u56f4\uff0c\u4e3b\u952e\u5b57\u6bb5\u3002"),
        (zh("\u4f18\u5148\u786e\u8ba4 2"), "3_EntityLinks_ToFill\uff1a\u786e\u8ba4\u8c01\u4f9d\u8d56\u8c01\uff0c\u8c01\u8c03\u7528\u8c01\uff0c\u65b9\u5411\u662f\u4e0d\u662f\u5bf9\u3002"),
        (zh("\u4f18\u5148\u786e\u8ba4 3"), "4_DataSets_ToFill \u548c 5_DataLinks_ToFill\uff1a\u786e\u8ba4\u8bc1\u636e\u7c7b\u578b\u548c\u5173\u8054\u5b57\u6bb5\u662f\u5426\u5408\u7406\u3002"),
        (zh("\u4f18\u5148\u786e\u8ba4 4"), "11_OpenQuestions\uff1a\u628a\u4ecd\u62ff\u4e0d\u51c6\u7684\u5730\u65b9\u8865\u6210\u660e\u786e\u7ed3\u8bba\u3002"),
        (
            zh("\u5b8c\u6210\u6807\u51c6"),
            zh("\u91cd\u70b9\u884c\u90fd\u586b\u4e86\u2018\u4e13\u5bb6\u5904\u7406\u7ed3\u679c\u2019\uff0c\u6709\u4fee\u6539\u7684\u5730\u65b9\u5199\u4e86\u2018\u4fee\u6539\u540e\u7684\u503c\u2019\u548c\u2018\u4fee\u6539/\u8865\u5145\u8bf4\u660e\u2019\uff0c\u771f\u7684\u4e0d\u786e\u5b9a\u7684\u5730\u65b9\u5199\u5230\u2018\u4e13\u5bb6\u5907\u6ce8/\u5f85\u786e\u8ba4\u2019\u3002"),
        ),
    ]
    for row in rows:
        ws.append(list(row))
    ws.freeze_panes = "A2"
    autosize(ws)


def main():
    if not SOURCE.exists():
        raise FileNotFoundError(f"Source workbook not found: {SOURCE}")

    src_wb = openpyxl.load_workbook(str(SOURCE))
    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active)

    build_guide_sheet(out_wb)

    for sheet_name in src_wb.sheetnames:
        src_ws = src_wb[sheet_name]
        ws = out_wb.create_sheet(sheet_name)
        copy_sheet_values(src_ws, ws)

        if sheet_name == "2_EntitySets_ToFill":
            highlight_columns(ws, ["display_name", zh("\u5bf9\u8c61\u8303\u56f4"), zh("\u5019\u9009\u4e3b\u952e"), zh("\u5019\u9009\u8f85\u52a9\u5b57\u6bb5"), zh("\u9884\u586b\u4f9d\u636e"), zh("\u8fd8\u9700\u8981\u8865\u4ec0\u4e48")])
            highlight_uncertain_cells(ws, [zh("\u5019\u9009\u4e3b\u952e"), zh("\u8fd8\u9700\u8981\u8865\u4ec0\u4e48"), zh("\u51b3\u7b56")])
            add_confirmation_columns(ws, lambda row: zh("\u8bf7\u91cd\u70b9\u786e\u8ba4\uff1a\u5bf9\u8c61\u4e2d\u6587\u540d\u3001\u5bf9\u8c61\u8303\u56f4\u3001\u5019\u9009\u4e3b\u952e\u662f\u5426\u6b63\u786e\uff1b\u5982\u679c\u4e3b\u5bf9\u8c61\u4e0d\u5bf9\u8bf7\u76f4\u63a5\u6539\u539f\u5355\u5143\u683c\u3002"))
        elif sheet_name == "3_EntityLinks_ToFill":
            highlight_columns(ws, ["src entity_set", "dest entity_set", "entity_link_type", zh("\u5019\u9009 fields_mapping"), zh("\u4f9d\u636e")])
            highlight_uncertain_cells(ws, [zh("\u5019\u9009 fields_mapping"), zh("\u8fd8\u9700\u8981\u8865\u4ec0\u4e48")])
            add_confirmation_columns(ws, lambda row: zh("\u8bf7\u91cd\u70b9\u786e\u8ba4\uff1a\u8fd9\u6761\u5173\u7cfb\u662f\u5426\u6210\u7acb\uff0c\u65b9\u5411\u662f\u5426\u6b63\u786e\uff0cfields_mapping \u662f\u5426\u5408\u7406\u3002"))
        elif sheet_name == "4_DataSets_ToFill":
            highlight_columns(ws, [zh("\u7528\u9014"), zh("\u6838\u5fc3\u5b57\u6bb5/\u6807\u7b7e"), zh("\u5173\u952e\u914d\u7f6e"), zh("\u5df2\u89c2\u6d4b\u5230\u7684 serviceName"), zh("\u8fd8\u9700\u8981\u8865\u4ec0\u4e48")])
            highlight_uncertain_cells(ws, [zh("\u5173\u952e\u914d\u7f6e"), zh("\u8fd8\u9700\u8981\u8865\u4ec0\u4e48")])
            add_confirmation_columns(ws, lambda row: zh("\u8bf7\u91cd\u70b9\u786e\u8ba4\uff1a\u8fd9\u79cd\u8bc1\u636e\u662f\u5426\u786e\u5b9e\u6709\u7528\uff0c\u5b57\u6bb5/\u6807\u7b7e\u662f\u5426\u591f\u6392\u969c\u4f7f\u7528\u3002"))
        elif sheet_name == "5_DataLinks_ToFill":
            highlight_columns(ws, ["src entity_set", "dest dataset", zh("\u5019\u9009 fields_mapping"), zh("\u72b6\u6001"), zh("\u5907\u6ce8")])
            highlight_uncertain_cells(ws, [zh("\u5019\u9009 fields_mapping"), zh("\u72b6\u6001"), zh("\u5907\u6ce8")])
            add_confirmation_columns(ws, lambda row: zh("\u8bf7\u91cd\u70b9\u786e\u8ba4\uff1a\u5b9e\u4f53\u548c\u8bc1\u636e\u4e4b\u95f4\u662f\u5426\u771f\u80fd\u9760\u8fd9\u4e9b\u5b57\u6bb5\u5173\u8054\u8d77\u6765\u3002"))
        elif sheet_name == "6_Storages_ToFill":
            highlight_columns(ws, ["spec.name", zh("\u5f53\u524d\u672c\u5730\u8def\u5f84"), zh("\u5019\u9009 properties"), zh("\u8fd8\u9700\u8981\u8865\u4ec0\u4e48")])
            highlight_uncertain_cells(ws, [zh("\u5019\u9009 properties"), zh("\u8fd8\u9700\u8981\u8865\u4ec0\u4e48")])
            add_confirmation_columns(ws, lambda row: zh("\u8bf7\u91cd\u70b9\u786e\u8ba4\uff1a\u771f\u5b9e\u67e5\u8be2\u5165\u53e3\u3001\u7d22\u5f15\u540d\u3001\u6570\u636e\u6e90\u540d\u662f\u5426\u9700\u8981\u8865\u5145\u3002"))
        elif sheet_name == "7_StorageLinks_ToFill":
            highlight_columns(ws, ["src dataset", "dest storage", zh("\u5019\u9009 fields_mapping"), zh("\u8fd8\u9700\u8981\u8865\u4ec0\u4e48")])
            highlight_uncertain_cells(ws, [zh("\u5019\u9009 fields_mapping"), zh("\u8fd8\u9700\u8981\u8865\u4ec0\u4e48")])
            add_confirmation_columns(ws, lambda row: zh("\u8bf7\u91cd\u70b9\u786e\u8ba4\uff1a\u8fd9\u4e2a\u6570\u636e\u96c6\u548c\u8fd9\u4e2a\u5b58\u50a8\u662f\u5426\u771f\u662f\u4e00\u4e00\u5bf9\u5e94\u3002"))
        elif sheet_name == "8_ObservedFieldProfile":
            highlight_columns(ws, ["Suggested Use", "Candidate Target", "Notes"])
            highlight_uncertain_cells(ws, ["Suggested Use", "Candidate Target", "Notes"])
            add_confirmation_columns(ws, lambda row: zh("\u5982\u5b57\u6bb5\u7528\u9014\u5224\u65ad\u4e0d\u5bf9\uff0c\u8bf7\u76f4\u63a5\u6539 Suggested Use / Candidate Target\u3002"))
        elif sheet_name == "9_ServiceInventory":
            highlight_columns(ws, ["Observed serviceName", "Guessed Object Type", "Expert Mapping To Fill"])
            highlight_uncertain_cells(ws, ["Guessed Object Type", "Expert Mapping To Fill"])
            add_confirmation_columns(ws, lambda row: zh("\u8bf7\u628a\u8fd9\u4e2a serviceName \u5bf9\u5e94\u5230\u4e13\u5bb6\u53e3\u5f84\u91cc\u7684\u5bf9\u8c61\u540d\uff1b\u5982\u679c\u770b\u4e0d\u51fa\u6765\u8bf7\u5907\u6ce8\u3002"))
        elif sheet_name == "10_ExpertRulesAndCases":
            highlight_columns(ws, ["Scene/Case", "Symptom", "Main Object", "Rule/Conclusion", "Common Root Cause", "Action"])
            highlight_uncertain_cells(ws, ["Main Object", "Rule/Conclusion", "Common Root Cause"])
            add_confirmation_columns(ws, lambda row: zh("\u8bf7\u91cd\u70b9\u786e\u8ba4\uff1a\u73b0\u8c61\u3001\u4e3b\u5bf9\u8c61\u3001\u5224\u65ad\u89c4\u5219\u3001\u5e38\u89c1\u6839\u56e0\u3001\u5904\u7406\u52a8\u4f5c\u662f\u5426\u51c6\u786e\u3002"))
        elif sheet_name == "11_OpenQuestions":
            highlight_columns(ws, ["Topic", "Question", "Why It Matters"])
            highlight_uncertain_cells(ws, ["Question"])
            add_confirmation_columns(ws, lambda row: zh("\u8bf7\u76f4\u63a5\u7ed9\u51fa\u7ed3\u8bba\uff1b\u5982\u679c\u6682\u65f6\u4e0d\u80fd\u5b9a\uff0c\u8bf7\u5199\u6e05\u695a\u8fd8\u7f3a\u4ec0\u4e48\u4fe1\u606f\u3002"))

        ws.row_dimensions[1].height = 38
        autosize(ws)

    out_wb.save(TARGET)
    print(TARGET)


if __name__ == "__main__":
    main()
